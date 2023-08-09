import mimetypes
import smtplib
from email.message import EmailMessage
from pathlib import Path
from time import sleep
from typing import Optional
from urllib.parse import parse_qs, urlparse

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from config import *


def send_email(
    file_path: list[str] | list[Path],
    email_addresses: Optional[list[str]] = None,
    subject: Optional[str] = None,
    body: Optional[str] = None,
):
    list_file_path = file_path
    message = EmailMessage()
    message["From"] = USER_MAIL
    if email_addresses:
        to = ", ".join(email_addresses)
    else:
        to = USER_MAIL
    message["To"] = to
    if subject:
        message["Subject"] = subject
    if body:
        message.add_alternative(body, subtype="html")
    for file_path_ in list_file_path:
        with open(file_path_, "rb") as f:
            ctype, encoding = mimetypes.guess_type(file_path_)
            if ctype is None or encoding is not None:
                ctype = "application/octet-stream"
            maintype, subtype = ctype.split("/", 1)
            message.add_attachment(
                f.read(),
                maintype=maintype,
                subtype=subtype,
                filename=file_path_.name,
            )
    session = smtplib.SMTP("smtp.gmail.com", 587)
    session.starttls()
    session.login(USER_MAIL, USER_PASS)
    # session.send_message(message)
    session.sendmail(USER_MAIL, to, message.as_string())
    session.quit()


def have_terms_banned_(list_terms_banned, name):
    have_terms_banned = False
    for word in list_terms_banned:
        if word in name:
            have_terms_banned = True

    return have_terms_banned


def have_all_terms_product_(list_terms_products, name):
    have_all_terms_product = True
    for word in list_terms_products:
        if word not in name:
            have_all_terms_product = False

    return have_all_terms_product


def formated_price(price):
    formated_price = (
        price.replace("R$", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )

    return formated_price


def search_google_shopping(
    driver, product, terms_banned, price_min, price_max
):
    # entrar no google
    driver.get("https://www.google.com/")
    driver.maximize_window()

    # tratar os valores que vieram da tabela
    product = product.lower()
    terms_banned = terms_banned.lower()
    list_terms_banned = terms_banned.split(" ")
    list_terms_products = product.split(" ")
    price_max = float(price_max)
    price_min = float(price_min)

    # pesquisar o nome do produto no google
    driver.find_element(
        By.XPATH,
        '//*[@id="APjFqb"]',
    ).send_keys(product)
    driver.find_element(
        By.XPATH,
        '//*[@id="APjFqb"]',
    ).send_keys(Keys.ENTER)

    sleep(2)
    # clicar na aba shopping
    elements = driver.find_elements(By.CLASS_NAME, "GKS7s")
    for item in elements:
        if "Shopping" in item.text:
            item.click()
            break

    # pegar a lista de resultados da busca no google shopping
    list_results = driver.find_elements(By.CLASS_NAME, "sh-dgr__grid-result")
    sleep(5)

    # para cada resultado, ele vai verificar se o resultado corresponde a todas as nossas condicoes
    list_offers = []  # lista que a função vai me dar como resposta
    for results in list_results:
        name = results.find_element(By.CLASS_NAME, "tAxDx").text
        name = name.lower()

        # verificacao do nome - se no nome tem algum termo banido
        have_terms_banned = have_terms_banned_(list_terms_banned, name)

        # verificar se no nome tem todos os termos do nome do produto
        have_all_terms_product = have_all_terms_product_(
            list_terms_products, name
        )

        if (
            not have_terms_banned and have_all_terms_product
        ):  # verificando o nome
            try:
                price = results.find_element(By.CLASS_NAME, "a8Pemb").text
                price = formated_price(price)
                price = float(price)
                # verificando se o preco ta dentro do minimo e maximo
                if price_min <= price <= price_max:
                    element_link = results.find_element(
                        By.CLASS_NAME, "aULzUe"
                    )
                    element_father = element_link.find_element(By.XPATH, "..")
                    link = element_father.get_attribute("href")
                    parsed = urlparse(link)
                    parsed = parse_qs(parsed.query)
                    if "url" in parsed:
                        link = parsed["url"][0]
                    list_offers.append((name, price, link))
            except:
                continue

    sleep(10)

    return list_offers


def search_buscape(driver, product, terms_banned, price_min, price_max):
    # tratar os valores da função
    price_max = float(price_max)
    price_min = float(price_min)
    product = product.lower()
    terms_banned = terms_banned.lower()
    list_terms_banned = terms_banned.split(" ")
    list_terms_products = product.split(" ")

    # entrar no buscape
    driver.get("https://www.buscape.com.br/")
    driver.maximize_window()

    # pesquisar pelo produto no buscape
    driver.find_element(
        By.XPATH,
        '//*[@id="new-header"]/div[1]/div/div/div[3]/div/div/div[2]/div/div[1]/input',
    ).send_keys(product, Keys.ENTER)
    sleep(10)

    # pegar a lista de resultados da busca do buscape
    sleep(5)
    list_results = driver.find_elements(
        By.CLASS_NAME, "SearchCard_ProductCard_Inner__7JhKb"
    )
    sleep(5)

    # para cada resultado
    list_offers = []
    for results in list_results:
        try:
            price = results.find_element(
                By.CLASS_NAME, "Text_MobileHeadingS__Zxam2"
            ).text
            name = results.find_element(
                By.CLASS_NAME, "SearchCard_ProductCard_Name__ZaO5o"
            ).text
            name = name.lower()
            link = results.get_attribute("href")

            # verificacao do nome - se no nome tem algum termo banido
            have_terms_banned = have_terms_banned_(list_terms_banned, name)

            # verificar se no nome tem todos os termos do nome do produto
            have_all_terms_product = have_all_terms_product_(
                list_terms_products, name
            )

            if not have_terms_banned and have_all_terms_product:
                price = formated_price(price)
                price = float(price)
                if price_min <= price <= price_max:
                    list_offers.append((name, price, link))
        except:
            pass

    sleep(10)

    return list_offers


def generate_styled_excel(table_offers):
    col_lens = (
        max(table_offers[c].apply(str).str.len().max() * 1.2, 8)
        for c in table_offers.columns
    )

    css_alt_rows = "background-color: powderblue; color: black;"
    css_indexes = "background-color: steelblue; color: white;"

    with pd.ExcelWriter("test/ofertas.xlsx") as writer:
        (
            table_offers.style.apply(
                lambda col: np.where(col.index % 2, css_alt_rows, None)
            )  # alternating rows
            .applymap_index(
                lambda _: css_indexes, axis=0
            )  # row indexes (pandas 1.4.0+)
            .applymap_index(
                lambda _: css_indexes, axis=1
            )  # col indexes (pandas 1.4.0+)
        ).to_excel(writer, sheet_name="list_offers", index=False)

        align = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws = writer.sheets["list_offers"]
        for i, col_len in enumerate(col_lens):
            ws.column_dimensions[get_column_letter(i + 1)].width = col_len

        ws.sheet_view.showGridLines = False

        for row in range(1, len(table_offers) + 2):
            for col in range(1, len(table_offers.columns) + 1):
                ws.cell(row, col).alignment = align
                ws.cell(row, col).border = border


def list_offers_found(table_products):
    driver = webdriver.Chrome()
    table_offers = list()

    for line in table_products.index:
        product = table_products.loc[line, "Nome"]
        terms_banned = table_products.loc[line, "Termos banidos"]
        price_min = table_products.loc[line, "Preço mínimo"]
        price_max = table_products.loc[line, "Preço máximo"]

        list_offers_google_shopping = search_google_shopping(
            driver, product, terms_banned, price_min, price_max
        )
        for offer in list_offers_google_shopping:
            table_google_shopping = dict(
                zip(
                    ["Produto", "Preço", "Link"],
                    offer,
                )
            )
            table_offers.append(table_google_shopping)

        list_offers_buscape = search_buscape(
            driver, product, terms_banned, price_min, price_max
        )
        for offer in list_offers_buscape:
            table_buscape = dict(zip(["Produto", "Preço", "Link"], offer))
            table_offers.append(table_buscape)

    # exportar pro excel
    table_offers = pd.DataFrame(table_offers)
    table_offers = table_offers.reset_index(drop=True)
    generate_styled_excel(table_offers)

    return table_offers


if __name__ == "__main__":
    driver = webdriver.Chrome()

    lista_google_shopping = search_google_shopping(
        driver, "iphone 12 64gb", "mini watch", "3000", "3500"
    )
    print(lista_google_shopping)

    lista_buscape = search_buscape(
        driver, "iphone 12 64gb", "mini watch", "3000", "3500"
    )
    print(lista_buscape)
